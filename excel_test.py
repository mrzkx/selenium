import copy
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class MyExcel():
    def __init__(self):
        self.initial_table_path = "软件生成文档/初始表格.xlsm"
        self.account_table_path = "自建品测试账号.xlsx"
        self.source_data_table_path = "总数据/总数据.xlsm"
        self.write_info_list = []
        self.path = "./shop_info_xlsm/"
        if not os.path.exists(self.path):
            os.mkdir(self.path)

    def get_excel_fonts(self, excel_name):
        """
        带格式复制
        """
        wb = load_workbook(self.initial_table_path)
        wb2 = Workbook()
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        sheet2 = wb2.create_sheet(sheet_name)
        # tab颜色
        sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor
        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                sheet2.merge_cells(cell2)
        for i, row in enumerate(sheet.iter_rows()):
            sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
            for j, cell in enumerate(row):
                sheet2.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[get_column_letter(j + 1)].width
                sheet2.cell(row=i + 1, column=j + 1, value=cell.value)
                # 设置单元格格式
                source_cell = sheet.cell(i + 1, j + 1)
                target_cell = sheet2.cell(i + 1, j + 1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(
                        source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)
        if 'Sheet' in wb2.sheetnames:
            del wb2['Sheet']
        wb.save(self.path + excel_name)
        wb.close()
        wb2.close()

    def get_excel_name(self):
        """
        获取新表名
        """
        excel_name_list = []
        wb = load_workbook(self.account_table_path)
        excel_id_sheet = wb["Sheet1"]
        excel_id_list = excel_id_sheet["B"]
        for item in excel_id_list[1:]:
            excel_name = "int_pdt_seller_import_%s.xlsm" % item.value
            excel_name_list.append(excel_name)
        wb.close()
        return excel_name_list

    def get_total_excel(self):
        """
        获取数据，存入到列表中
        """

        rows_cases_list = []
        wb = load_workbook(self.source_data_table_path)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        start_num = 4
        end_num = sheet.max_row + 1
        for item in range(start_num, end_num):
            rows_cases_list.append(sheet[item])
        self.write_info_list.append(rows_cases_list)
        wb.close()
        return rows_cases_list

    def write_excel(self, excel_name, rows_cases_list):
        """
        写入表
        """

        wb = load_workbook(excel_name)
        ws = wb.active
        for row_index, row_item in enumerate(rows_cases_list):
            for col_index, col_item in enumerate(row_item):
                # exit()
                # 写入
                ws.cell(row=row_index + 4, column=col_index + 1, value=col_item.value)
        wb.save(self.path + excel_name)
        wb.close()
        return "OK"


if __name__ == "__main__":
    excel = MyExcel()
    excel_name_list = excel.get_excel_name()
    rows_cases_list = excel.get_total_excel()
    for excel_name in excel_name_list:
        rows_cases = rows_cases_list[0:100]
        excel.get_excel_fonts(excel_name)
        print("{}已初始化！".format(excel_name))
        excel.write_excel(excel_name, rows_cases)
        print("{}写入成功！".format(excel_name))
        for item in rows_cases:
            rows_cases_list.remove(item)
